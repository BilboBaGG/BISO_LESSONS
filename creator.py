from openpyxl import load_workbook # EXCEL

NAME = "БИСО-01-22"

# Обработка расписания из файла
def get_lessons():

	lessons = [["" for i in range(6)] for i in range(12)] # 6 пар каждый день в течение двух недель

	excel_file = load_workbook('./1-kurs-IKB.xlsx')

	sheet_names = excel_file.sheetnames

	for sheet_name in sheet_names:

		sheet = excel_file[sheet_name]

		if sheet["K2"].value == NAME or sheet["F2"].value == NAME: # Поиск группы

			if sheet["F2"].value == NAME:
				row = "F"
			else:
				row = "K"

			for i in range(4,76): # Строки таблицы с 4-ой по 75-ую

				# В четных ячейках лежат пары нечетных недель и наоборот

				day = (i-4)//12 + (i%2)*6 # День пары
				pare_index = ((i-4)%12)//2 # Номер пары 
				


				if sheet[row+str(i)].value != "": # Если пара существует

					lessons[day][pare_index] = [sheet[row+str(i)].value,sheet[chr(ord(row)+1)+str(i)].value,sheet[chr(ord(row)+2)+str(i)].value,sheet[chr(ord(row)+3)+str(i)].value]
					

					# Делаем вывод красивым и понятным


					if len(lessons[day][pare_index][0].split("\n"))>1: # удаление повторов и разъединение двойных пар

						for i in range(0,4):
							if lessons[day][pare_index][i].split("\n")[0]!=lessons[day][pare_index][i].split("\n")[1]:
								lessons[day][pare_index][i] = " или ".join(lessons[day][pare_index][i].split("\n"))
							else:
								lessons[day][pare_index][i] = lessons[day][pare_index][i].split('\n')[0]


					# Удаление доп иформации о паре

					if "кр." in lessons[day][pare_index][0]:
						lessons[day][pare_index][0] = " ".join(lessons[day][pare_index][0].split(" ")[3:])

					if ord(lessons[day][pare_index][0][0]) < 1040:

						lessons[day][pare_index][0] = " ".join(lessons[day][pare_index][0].split(" ")[2:])

					# Обработка здания
					place = lessons[day][pare_index][3].split(" ")
					for i in range(len(place)):
						if place[i] == "(С-20)":
							place[i] = "на Стромынке"
						elif place[i] == "(В-78)":
							place[i] = "на Вернадке 78"
						elif place[i] == "(МП-1)":
							place[i] = "на Пироговке"
					lessons[day][pare_index][3] = " ".join(place)

	# Добавляем воскресения, чтобы можно было работать с датами в семидневных неделях			
	lessons.insert(6,[])
	lessons.append([])
	#print(lessons)
	return lessons

"""lessons = get_lessons()

for i in range(len(lessons)):
	print(i+1)
	print()
	for j in range(len(lessons[i])):
		if type(lessons[i][j]) != type(" "):
			print(j+1,lessons[i][j])
"""